from tkinter import filedialog, messagebox
from pdf_modifications import TwoPixelReportGenerator, EightPixelReportGenerator
import random
import pandas as pd
import re
from openpyxl import load_workbook
import tkinter as tk
from pathlib import Path
from PIL import Image, ImageTk
from tkcalendar import DateEntry
import os
from collections import defaultdict

class SolarAnalysisApp:
    """GUI application for processing and 
    analyzing photovoltaic solar measurement data.
    
    Responsibilities:
    - Provide a user interface for selecting 2-pixel Excel files or 8-pixel data folders.
    - Collect experiment metadata (experimenter name, date, scan rate, sun intensity, temperature).
    - Validate and check the structure of selected Excel templates.
    - Orchestrate generation of PDF reports for both 2-pixel and 8-pixel devices.
    - Manage application state by showing/hiding UI sections based on user selection.
    - Display status messages and errors via dialog boxes for a clear user experience.
    """

    def __init__(self, root):

        # Create the main window
        self.root = root
        self.root.title("Reporting Utility for Solar Logging and Analysis of Numerical data PRO")

        # StringVar for tracking the selected 2-pixel Excel file path
        self.selected_file = tk.StringVar()

        # Holds the base folder path for 8-pixel data (set after folder selection)
        self.selected_8_pixel_folder = None

        self.jv_file_paths = {}  # Dictionary to store JV file paths for 8-pixel analysis

        # Build all UI sections (logo, welcome text, selection buttons, data entry, file/folder pickers, action buttons, footer)
        self._setup_logo() # Setup top banner and description
        self._setup_welcome_section()
        self._setup_selection_buttons() # Setup device‐type selection controls
        self._setup_experimenter_info() # Setup metadata input sections (experimenter name, date, scan parameters)
        self._setup_date_section()
        self._setup_metadata_section()
        self._setup_file_section_2_pixel() # Setup file chooser for 2-pixel mode
        self._setup_action_buttons() # Setup action buttons (Generate/Check)
        self._setup_footer() # Setup footer
        self._setup_folder_section_8_pixel() # Setup folder chooser for 8-pixel mode

        # Hide UI sections until a device type (2-pixel or 8-pixel) is selected
        self.hide_all_sections()

    def _setup_logo(self):
        """Setup the application logo"""

        self.base_dir = Path(__file__).parent.parent
        logo_path = self.base_dir / "Images/smat_logo_png.png" # Constructing the path of the logo 
        
        # Load and display application logo; fallback to console warning if missing
        try:
            icon = Image.open(logo_path)
            icon = icon.resize((75, 96), Image.Resampling.LANCZOS)
            icon = ImageTk.PhotoImage(icon)
            icon_label = tk.Label(self.root, image=icon)
            icon_label.image = icon  # Keep a reference
            icon_label.pack()

        except FileNotFoundError:
            print("Logo icon not found.")

    def _setup_welcome_section(self):
        """Setup welcome message and description"""

        welcome_label = tk.Label(
            self.root, 
            text="Welcome to R.U.S.L.A.N. PRO", 
            font=("Arial", 14, "bold")
        )
        welcome_label.pack(pady=10)

        description = """
        This program processes Excel files of photovoltaic solar measurements.
        Please select the type of solar cell for analysis.
        """
        desc_label = tk.Label(
            self.root, 
            text=description, 
            justify="center", 
            font=("Arial", 10), 
            padx=10, 
            pady=5
        )
        desc_label.pack()

    def _setup_selection_buttons(self):
        """Setup the 2-pixel and 8-pixel selection buttons"""

        # Create frame to hold pixel selection buttons
        self.selection_frame = tk.Frame(self.root)
        self.selection_frame.pack(pady=5)

        # Define button for 2-pixel solar cell analysis
        two_pixel_button = tk.Button(
            self.selection_frame, 
            text="2-Pixel", 
            command=self.show_2_pixel_section, 
            font=("Arial", 12, "bold"),
            width=8,
            activeforeground="white",
            activebackground="black",
            borderwidth=4,
            relief="raised",
            cursor="hand2"
        )
        two_pixel_button.pack(side="left", padx=20, pady=10)

        # Define button for 8-pixel solar cell analysis
        eight_pixel_button = tk.Button(
            self.selection_frame, 
            text="8-Pixel", 
            command=self.show_8_pixel_section, 
            font=("Arial", 12, "bold"),
            width=8,
            activeforeground="white",
            activebackground="black",
            borderwidth=4,
            relief="raised",
            cursor="hand2"
        )
        eight_pixel_button.pack(side="left", padx=20, pady=10)

    def _setup_experimenter_info(self):
        """Setup experimenter name entry"""

        # Create frame and entry widget for experimenter name
        self.name_frame = tk.Frame(self.root)
        self.name_frame.pack(pady=10)

        name_label = tk.Label(
            self.name_frame, 
            text="Experimenter Name:", 
            font=("Arial", 10)
        )
        name_label.pack(side="left", padx=5)

        self.name_entry = tk.Entry(
            self.name_frame, 
            font=("Arial", 10), 
            width=25
        )
        self.name_entry.pack(side="left")
    
    def _setup_date_section(self):
        """Setup experiment date section"""

        # Create frame and date picker for measurement date
        self.date_frame = tk.Frame(self.root)
        tk.Label(
            self.date_frame, 
            text="Measurement Date (DD/MM/YYYY):", 
            font=("Arial", 10)
        ).pack(side="left", padx=5)

        self.date_entry = DateEntry(
            self.date_frame, 
            date_pattern='dd/mm/yyyy', 
            font=("Arial", 10), 
            width=12
        )
        self.date_entry.pack(side="left", padx=5)

    def _setup_metadata_section(self):
        """Setup metadata entries (scan rate, sun intensity, temperature)"""
        
        # Create frame and entry widgets for measurement metadata
        self.metadata_frame = tk.Frame(self.root)

        # Scan Rate input
        tk.Label(
            self.metadata_frame, 
            text="Scan Rate (V/s):", 
            font=("Arial", 10)
        ).grid(row=0, column=0, padx=5)
        self.scan_rate_entry = tk.Entry(
            self.metadata_frame, 
            font=("Arial", 10), 
            width=10
        )
        self.scan_rate_entry.grid(row=0, column=1, padx=5)

        # Sun Intensity input
        tk.Label(
            self.metadata_frame, 
            text="Sun Intensity (W/m²):", 
            font=("Arial", 10)
        ).grid(row=1, column=0, padx=5)
        self.sun_intensity_entry = tk.Entry(
            self.metadata_frame, 
            font=("Arial", 10), 
            width=10
        )
        self.sun_intensity_entry.grid(row=1, column=1, padx=5)

        # Temperature input
        tk.Label(
            self.metadata_frame, 
            text="Temperature (°C):", 
            font=("Arial", 10)
        ).grid(row=2, column=0, padx=5)
        self.temp_entry = tk.Entry(
            self.metadata_frame, 
            font=("Arial", 10), 
            width=10
        )
        self.temp_entry.grid(row=2, column=1, padx=5)

        self.metadata_frame.pack(pady=10)
    
    def _setup_file_section_2_pixel(self):
        """Setup file selection sections for both 2-pixel and 8-pixel modes"""

        # Create frame for 2-pixel file selection
        self.file_selection_frame  = tk.Frame(self.root)

        # StringVar to track the selected file path
        self.selected_file = tk.StringVar()
        self.file_label = tk.Label(
            self.file_selection_frame , 
            text="No file selected", 
            font=("Arial", 10), 
            fg="red"
        )

        # Label to display selected file status
        self.file_label.pack(pady=5)

        # Button to trigger file chooser dialog
        choose_file_button  = tk.Button(
            self.file_selection_frame , 
            text="Choose File", 
            command=lambda: self.choose_file()
        )
        choose_file_button.pack(pady=5)
        
    def _setup_folder_section_8_pixel(self):
        """Setup folder selection section for 8-pixel mode"""
        
        # Create frame for 8-pixel folder selection
        self.folder_selection_frame = tk.Frame(self.root)

        # StringVar to track the selected folder path
        self.selected_folder = tk.StringVar()
        self.selected_folder.set("No folder selected")

        # Label that dynamically shows selected folder path
        self.folder_label = tk.Label(
            self.folder_selection_frame,
            textvariable=self.selected_folder,
            font=("Arial", 10),
            fg="red"
        )
        self.folder_label.pack(pady=5)

        # Button to open folder chooser dialog for 8-pixel measurements
        choose_folder_button = tk.Button(
            self.folder_selection_frame,
            text="Choose Folder",
            command=lambda: self.choose_8_pixel_folder()
        )
        choose_folder_button.pack(pady=5)
    
    def choose_8_pixel_folder(self):
        """Prompt the user to select the base folder for 8-pixel data"""

        # Open folder selection dialog for user
        folder_path = filedialog.askdirectory(title="Select Base Folder for 8-Pixel Measurements")
        
        # If folder selected, update label and collect sample information
        if folder_path:
            self.selected_8_pixel_folder = folder_path
            self.selected_folder.set(f"Selected 8-pixel folder:\n{folder_path}")
            self.folder_label.config(fg="green")  # Change text color to green
            self.extract_sample_names()
            self.collect_jv_file_paths()  # <- collect CSV file paths
        
        else: # If no folder selected, reset UI and internal variables
            self.selected_folder.set("No folder selected")
            self.selected_8_pixel_folder = None
            self.folder_label.config(fg="red")  # Reset text color to red (or any color you want)
     
    def extract_sample_names(self):
        """Extract unique sample names and pixel counts from folder names like 'sample X[1]' """

        # Ensure a folder has been selected before proceeding
        if not hasattr(self, 'selected_8_pixel_folder') or not self.selected_8_pixel_folder:
            print("[ERROR] No folder selected for 8-pixel analysis. Operation aborted.")
            return

        # List all folders in the selected base directory
        folder_names = sorted(os.listdir(self.selected_8_pixel_folder))

        # Regex pattern to match folders with format: 'SampleName[PixelNumber]'
        sample_pixel_pattern = re.compile(r'^(.*)\[(\d+)\]$')  # Match 'sample name[1]'... 'sample name[8]'

        # Dictionary to count number of pixels per sample
        sample_counter = defaultdict(int)
        
        # Iterate through folders and extract sample names and pixel counts
        for folder in folder_names:
            full_path = os.path.join(self.selected_8_pixel_folder, folder)
            if os.path.isdir(full_path):
                match = sample_pixel_pattern.match(folder)
                if match:
                    sample_name = match.group(1).strip()
                    sample_counter[sample_name] += 1
        
        # Handle case where no valid sample folders were found
        if not sample_counter:
            # No matching folders found – update GUI and abort
            self.selected_folder.set("No valid sample folders found (e.g., 'sample X[1]') in the selected directory.")
            self.folder_label.config(fg="red")
            self.sample_pixel_counts  = ()
            print(f"[ERROR] No valid sample folders found in {self.selected_8_pixel_folder}. Expected folder format: 'SampleName[PixelNumber]'.")

            return
        
        # Save extracted sample information as a tuple of (sample name, pixel count)
        self.sample_pixel_counts  = tuple((name, count) for name, count in sorted(sample_counter.items()))
        print(f"Sample info: {self.sample_pixel_counts }")       
          
    def collect_jv_file_paths(self):
        """Collect all CSV file paths ending with 'Perform parallel JV.csv' from each sample's pixel folders."""
        
        # Helper function to apply natural sorting (e.g., 1,2,10 instead of 1,10,2)
        def natural_sort_key(s):
            """Natural sort key function for sorting strings with numbers."""
            return [int(text) if text.isdigit() else text.lower() 
                    for text in re.split('([0-9]+)', s)] 

        # Ensure sample information and base folder are available before proceeding
        if not hasattr(self, 'sample_pixel_counts') or not self.sample_pixel_counts:
            print(f"[ERROR] No sample information extracted from selected base folder: {self.selected_8_pixel_folder}")
            return

        if not hasattr(self, 'selected_8_pixel_folder') or not self.selected_8_pixel_folder:
            print("[ERROR] No folder selected for 8-pixel analysis. Operation aborted.")
            return

        self.jv_file_paths = {} # Dictionary to map each sample to its pixel JV file paths

        self.target_jv_filename = "Perform parallel JV.csv"

        # Iterate over each sample and its pixel count
        for sample_name, pixel_count in self.sample_pixel_counts:
            pixel_files = {}

            for pixel_num in range(1, pixel_count + 1):
                # Build full path to each pixel folder (e.g., 'SampleX[1]')
                folder_name = f"{sample_name}[{pixel_num}]"
                folder_path = os.path.join(self.selected_8_pixel_folder, folder_name)
                folder_path = os.path.normpath(folder_path)

                # Check if pixel folder exists
                if not os.path.isdir(folder_path):
                    print(f"[WARNING] Folder not found: {folder_path}. Skipping this pixel.")
                    continue

                matched_jv_csv_files  = []

                try:
                    # Find all CSV files ending with 'Perform parallel JV.csv' in pixel folder
                    for file_name in sorted(os.listdir(folder_path), key=natural_sort_key):
                        if file_name.endswith(self.target_jv_filename):
                            matched_jv_csv_files .append(os.path.join(folder_path, file_name))
                except Exception as e:
                    messagebox.showerror(f"Error accessing folder: {folder_path}. Error: {e}")
                    print(f"[EXCEPTION] Failed to list files in {folder_path}. Exception: {e}")
                    continue
                
                # Map pixel number to list of matched JV files
                if matched_jv_csv_files :
                    pixel_files[pixel_num] = matched_jv_csv_files
                else:
                    print(f"[INFO] No JV files ending with 'Perform parallel JV.csv' found in {folder_path}. Pixel {pixel_num} skipped.")

            # After processing all pixels, assign to the main dictionary
            self.jv_file_paths[sample_name] = pixel_files

    def _setup_action_buttons(self):
        """Setup Generate Report, Check Template buttons for 2-Pixel data and Generate report button for 8-Pixel data"""
        self.action_buttons_frame  = tk.Frame(self.root)
        
        self.generate_2_pixel_button = tk.Button(
            self.action_buttons_frame , 
            text="Generate Report", 
            command=lambda: self.generate_report_2_pixel(),
            fg="white", 
            bg="blue"
        )
        self.generate_2_pixel_button.grid(row=0, column=0, padx=5)

        self.check_template_button = tk.Button(
            self.action_buttons_frame , 
            text="Check Template", 
            command=lambda: self.check_template(),
            fg="white", 
            bg="green"
        )
        self.check_template_button.grid(row=0, column=1, padx=5)

        # 8-Pixel generate button
        self.generate_8_pixel_button = tk.Button(
            self.root, 
            text="Generate Report", 
            command=lambda: self.generate_report_8_pixel(),
            fg="white", 
            bg="blue"
        )
    
    def _setup_footer(self):
        """Setup footer with institutional details"""
        footer_text = "Developed by the SMAT Research Group, University of Turku. © 2025 All Rights Reserved."
        footer_label = tk.Label(
            self.root, 
            text=footer_text, 
            font=("Arial", 10), 
            fg="gray"
        )
        footer_label.pack(side="bottom", pady=5)

        footer_button_frame = tk.Frame(self.root)

        # Website Button
        website_button = tk.Button(
            footer_button_frame, 
            text="Visit SMAT Research Group", 
            fg="blue", 
            cursor="hand2", 
            command=lambda: self.open_website()
        )
        website_button.grid(row=0, column=0, padx=5)

        # Instructions Button placed on the right side of the website button
        instructions_button = tk.Button(
            footer_button_frame, 
            text="Open Instructions",
            fg="orange", 
            cursor="hand2", 
            command=self.open_instructions
        )
        instructions_button.grid(row=0, column=1, padx=5)

        footer_button_frame.pack(side="bottom", pady=5)

    def show_2_pixel_section(self):
        """Show 2-pixel related sections"""
        self.hide_all_sections()
        self.date_frame.pack(pady=5)
        self.metadata_frame.pack(pady=5)
        self.file_selection_frame.pack(pady=5)
        self.action_buttons_frame .pack(pady=10)

    def show_8_pixel_section(self):
        """Show 8-pixel related sections"""
        self.hide_all_sections()
        self.folder_selection_frame.pack(pady=5)
        self.generate_8_pixel_button.pack(pady=10)

    def hide_all_sections(self):
        """Hide all optional sections"""
        self.date_frame.pack_forget()
        self.metadata_frame.pack_forget()
        self.file_selection_frame.pack_forget()
        self.action_buttons_frame.pack_forget()
        self.generate_8_pixel_button.pack_forget()
        self.folder_selection_frame.pack_forget()

    def choose_file(self):
        '''Function to handle file selection for 2-Pixel cells (does NOT generate report automatically)'''
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if file_path:
            self.selected_file.set(file_path)  # Store the selected file path
            self.file_label.config(text=f"Selected: {file_path.split('/')[-1]}", fg="green")  # Update UI

    def check_template(self):
        '''Function to check the template of the Excel file'''
        file_path = self.selected_file.get()
        
        if not file_path:
            messagebox.showwarning("Warning", "Please select an Excel file first.")
            return
        
        try:
            # Load the workbook with openpyxl (this preserves formulas)
            wb = load_workbook(file_path, data_only=False)
            excel_file = pd.ExcelFile(file_path)
            sheet_names = wb.sheetnames

            # Dynamically Generate Expected Sheet Names Based on Existing Sheets
            expected_order = []
            sheet_count = len(sheet_names) // 4  # Assuming sheets are grouped in sets of 4 (l-fw, l-rv, r-fw, r-rv)

            for num in range(1, sheet_count + 1):
                for side in ['l', 'r']:
                    for cycle in ['fw', 'rv']:
                        expected_order.append(f"{num}-{side}-{cycle}")

            # Step 1: Check Sheet Name Format and Order
            missing_sheets = set(expected_order) - set(sheet_names)
            unexpected_sheets = set(sheet_names) - set(expected_order)

            if missing_sheets:
                messagebox.showerror("Error", f"Missing sheets:\n" + "\n".join(sorted(missing_sheets)))
                return

            if unexpected_sheets:
                messagebox.showerror("Error", f"Unexpected sheets found:\n" + "\n".join(sorted(unexpected_sheets)))
                return
            
            sheet_errors = []

            for i, sheet in enumerate(sheet_names):
                if not re.match(r"^\d+-(l|r)-(rv|fw)$", sheet):  # Check format
                    sheet_errors.append(f"Invalid format: {sheet}")
                elif sheet != expected_order[i]:  # Check order
                    sheet_errors.append(f"Expected: {expected_order[i]}, but found: {sheet}")

            if sheet_errors:
                messagebox.showerror("Error", f"The Excel file has an issue in sheet names:\n" + "\n".join(sheet_errors))
                return
            
            # Required columns
            required_columns = ['VOLTAGE (V)', 'CURRENT (uA)', 'POWER (uW)', 
                                'VOLTAGE (V).1', 'mA/cm²', 'mW/cm²']
            
            # Dictionary to store missing columns for each sheet
            missing_columns_dict = {}
            
            # Step 2: Check Column Names in Each Sheet
            for sheet in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet)
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    missing_columns_dict[sheet] = missing_columns

            # If there are any missing columns, display an error message
            if missing_columns_dict:
                error_msg = "There are missing columns in some sheets:\n"
                for sheet, cols in missing_columns_dict.items():
                    error_msg += f"- {sheet} misses: {', '.join(cols)}\n"
                messagebox.showerror("Error", error_msg)
                return

            # Step 3: Check Formulas
            formula_errors = []

            for sheet in sheet_names:
                sheet_obj = wb[sheet]
                
                # Iterate over rows (starting from row 2 to skip headers)
                for i in range(2, sheet_obj.max_row + 1):

                    # Check formulas in specific columns
                    voltage_formula = sheet_obj[f'I{i}'].value  # Formula in I column
                    mA_formula = sheet_obj[f'J{i}'].value  # Formula in J column
                    mW_formula = sheet_obj[f'L{i}'].value  # Formula in L column

                    # Check if the formulas match the expected ones
                    if voltage_formula != f"=IF(B{i}<>\"\",B{i}*-1,#N/A)":
                        formula_errors.append(f"{sheet}: The 2nd 'VOLTAGE (V)' column missing formula in row {i}")
                    
                    if mA_formula != f"=IF(C{i}<>\"\",(C{i})/(1000*$N$4),#N/A)":
                        formula_errors.append(f"{sheet}: 'mA/cm²' missing formula in row {i}")
                    
                    if mW_formula != f"=IF(D{i}<>\"\",(D{i}*-1)/(1000*$N$4),#N/A)":
                        formula_errors.append(f"{sheet}: 'mW/cm²' missing formula in row {i}")

            if formula_errors:
                messagebox.showerror("Error", "\n".join(formula_errors))
                return

            messagebox.showinfo("Success", "The Excel file meets the template requirements! Great job – ready to generate reports!")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def generate_report_2_pixel(self):
        '''Generate report button for 2-Pixel solar cells'''

        quotes = [
        "Generating report... 🌞 'Solar research: because the sun won’t patent itself… yet.' ☀️",
        "Processing data... ⚡ 'More efficiency, less resistance—applies to both solar cells and life!' 🔋",
        "Compiling results... 📊 'Every great discovery starts with a bad simulation. Keep going!' 🚀",
        "Analyzing measurements... 🌎 'Solar cells don’t give up on cloudy days, and neither should you!' ☁️",
        "Optimizing performance... ☀️ 'Solar power: because fusion reactors in the sky are better than coal mines underground.' 🔥",
        "Simulating efficiency... 🔬 'Photons are basically tiny energy delivery guys, and you’re the boss deciding where they go!' 🚀",
        "Crunching numbers... 📈 'Efficiency isn’t just a goal, it’s a way of life—especially when working at 2 AM.' 🧑‍💻",
        "Refining calculations... 🌟 'Nature already made a fusion reactor. You just need to harness it!' 🌞",
        "Validating results... 🔍 'Solar panels work best in the sun. So do researchers with coffee.' ☕",
        "Finalizing report... 📝 'If solar cells can convert chaos into power, so can you!' 💪"]

        quote = random.choice(quotes)

        experiment_date = self.date_entry.get()
        experimenter = self.name_entry.get()
        file_path = self.selected_file.get()
        scan_rate = self.scan_rate_entry.get()
        sun_intensity = self.sun_intensity_entry.get()
        temperature = self.temp_entry.get()

        if not self.selected_file.get():
            messagebox.showwarning("Warning", "Please select an Excel file before generating the report.")
            return
        if not self.name_entry.get():
            messagebox.showwarning("Warning", "Please enter the experimenter's name.")
            return
        
        metadata = {
            "Experiment Date": experiment_date if experiment_date else "N/A",
            "Scan Rate": scan_rate if scan_rate else "N/A",
            "Sun Intensity": sun_intensity if sun_intensity else "N/A",
            "Temperature": temperature if temperature else "N/A"
        }

        messagebox.showinfo("Generating Report", f"{quote}\n\nGenerating report for {experimenter} using file {file_path}")

        report_generator = TwoPixelReportGenerator(file_path, experimenter, metadata)
        report_generator.save_pdf()  # Call the save_pdf method to generate the report

    def generate_report_8_pixel(self):
        '''Generate report button for 2-Pixel solar cells'''

        quotes = [
        "Generating report... 🌞 'Solar research: because the sun won’t patent itself… yet.' ☀️",
        "Processing data... ⚡ 'More efficiency, less resistance—applies to both solar cells and life!' 🔋",
        "Compiling results... 📊 'Every great discovery starts with a bad simulation. Keep going!' 🚀",
        "Analyzing measurements... 🌎 'Solar cells don’t give up on cloudy days, and neither should you!' ☁️",
        "Optimizing performance... ☀️ 'Solar power: because fusion reactors in the sky are better than coal mines underground.' 🔥",
        "Simulating efficiency... 🔬 'Photons are basically tiny energy delivery guys, and you’re the boss deciding where they go!' 🚀",
        "Crunching numbers... 📈 'Efficiency isn’t just a goal, it’s a way of life—especially when working at 2 AM.' 🧑‍💻",
        "Refining calculations... 🌟 'Nature already made a fusion reactor. You just need to harness it!' 🌞",
        "Validating results... 🔍 'Solar panels work best in the sun. So do researchers with coffee.' ☕",
        "Finalizing report... 📝 'If solar cells can convert chaos into power, so can you!' 💪"]

        quote = random.choice(quotes)

        experimenter = self.name_entry.get()

        if not self.selected_8_pixel_folder:
            messagebox.showwarning("Warning", "Please select the directory that includes the 8-Pixel samples before generating the report.")
            return
        
        if not self.name_entry.get():
            messagebox.showwarning("Warning", "Please enter the experimenter's name.")
            return

        messagebox.showinfo("Generating Report", f"{quote}\n\nGenerating report for {experimenter} using file directory {self.selected_8_pixel_folder}")

        report_generator = EightPixelReportGenerator(self.jv_file_paths, experimenter)
        report_generator.save_pdf()  # Call the save_pdf method to generate the report

    def open_website(self):
        '''Add a clickable website link'''
        import webbrowser
        webbrowser.open("https://sites.utu.fi/smat/")

    def open_instructions(self):
        """Open the instructions PDF"""

        data_file = os.path.join(self.base_dir, 'Instructions.pdf')
        if os.path.exists(data_file):
            os.startfile(data_file)  # This works only in Windows
        else:
            print(f"Error: {data_file} not found")